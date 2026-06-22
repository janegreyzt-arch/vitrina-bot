from __future__ import annotations

import asyncio
import calendar
import logging
import os
import re
import shutil
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import aiosqlite
from aiogram import Bot, Dispatcher, F
from aiogram.filters import BaseFilter, Command
from aiogram.types import (
    CallbackQuery,
    FSInputFile,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ====================== НАСТРОЙКИ ======================
def load_env(path: str = ".env") -> None:
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            os.environ.setdefault(key.strip(), value.strip())


load_env()
TOKEN = (
    os.environ.get("BOT_TOKEN")
    or os.environ.get("TELEGRAM_BOT_TOKEN")
    or os.environ.get("TOKEN")
)
if not TOKEN:
    raise RuntimeError(
        "BOT_TOKEN не задан. Добавьте переменную BOT_TOKEN в .env или в панели хостинга"
    )

_work_chat_raw = os.environ.get("WORK_CHAT_ID", "-1003052957786")
try:
    WORK_CHAT_ID = int(_work_chat_raw)
except ValueError as exc:
    raise RuntimeError(
        f"WORK_CHAT_ID должен быть числом, получено: {_work_chat_raw!r}"
    ) from exc

# На bothost данные в ./data сохраняются между redeploy (см. bothost.ru/docs).
DB_DIR = os.environ.get("DB_DIR", "data")
DB_PATH = os.environ.get("DATABASE_PATH") or os.path.join(DB_DIR, "vitrina_bot.db")
BACKUP_DIR = os.path.join(DB_DIR, "backups")
LEGACY_DB_PATH = "vitrina_bot.db"
BACKUP_KEEP_LOCAL = 14

MSK = ZoneInfo("Europe/Moscow")
POINTS = ["МН", "СМ", "Д1", "СОК", "ПТ", "ПК", "МСТИЛЬ"]
TIMES = ["10:00", "16:00", "20:00"]
SLOT_WINDOWS = {
    "10:00": ("09:30", "10:30"),
    "16:00": ("15:30", "16:30"),
    "20:00": ("19:30", "20:30"),
}
SLOT_END = {slot: window[1] for slot, window in SLOT_WINDOWS.items()}
NO_EVENING_POINTS = ["МСТИЛЬ", "СОК"]
WEEKEND_CLOSED_POINTS = ["МСТИЛЬ"]
WEEKEND_NO_EVENING_POINTS = ["СОК", "МН", "ПК", "СМ"]

ADMIN_CODE = "1506"
FINE_AMOUNT = 500

ARRIVAL_KEYWORDS = (
    "ПРИШЕЛ", "ПРИШЛА", "ПРИХОД", "ПРИБЫЛ", "ПРИБЫЛА",
    "НА ТОЧКЕ", "ЗАШЕЛ", "ЗАШЛА", "НАЧАЛ", "НА РАБОТЕ",
    "ВЕРНУЛСЯ", "ВЕРНУЛАСЬ", "ВЕРНУЛСЬ", "ВЕРНУЛС",
)
DEPARTURE_KEYWORDS = (
    "ВЫШЕЛ", "ВЫШЛА", "УШЕЛ", "УШЛА", "ЗАКОНЧИЛ", "КОНЕЦ", "УХОД",
    "ВЫХОД", "С ТОЧКИ", "СО СМЕНЫ", "СМЕНА ЗАКОНЧ",
    "ОТОШЕЛ", "ОТОШОЛ", "ОТОШЛА", "ОТОШ",
)
RETURN_KEYWORDS = ("ВЕРНУЛСЯ", "ВЕРНУЛАСЬ", "ВЕРНУЛСЬ", "ВЕРНУЛС")
STEP_OUT_KEYWORDS = ("ОТОШЕЛ", "ОТОШОЛ", "ОТОШЛА", "ОТОШ")

MONTH_NAMES = [
    "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]

# user_id -> {"dates": set[str], "year": int, "month": int}
admin_sessions: dict[int, dict] = {}
authenticated_admins: set[int] = set()

# ====================== БОТ ======================
bot = Bot(token=TOKEN, timeout=60)
dp = Dispatcher()
logger = logging.getLogger(__name__)

# ====================== ВРЕМЯ (МСК) ======================
def now_msk() -> datetime:
    return datetime.now(MSK)


def get_today() -> str:
    return now_msk().strftime("%Y-%m-%d")


def parse_date(s: str) -> datetime:
    return datetime.strptime(s, "%Y-%m-%d").replace(tzinfo=MSK)


def is_weekend(date: str) -> bool:
    return parse_date(date).weekday() >= 5


def is_vitrina_required(date: str, point: str, slot: str) -> bool:
    if is_weekend(date) and point in WEEKEND_CLOSED_POINTS:
        return False
    if slot == "20:00":
        if point in NO_EVENING_POINTS:
            return False
        if is_weekend(date) and point in WEEKEND_NO_EVENING_POINTS:
            return False
    return True


def required_slots_for_date(date: str, point: str) -> list[str]:
    return [slot for slot in TIMES if is_vitrina_required(date, point, slot)]


def get_current_time_slot(now: datetime | None = None) -> str:
    now = now or now_msk()
    current = now.strftime("%H:%M")

    for slot, (start, end) in SLOT_WINDOWS.items():
        if start <= current <= end:
            return slot

    if current < SLOT_WINDOWS["10:00"][0]:
        return "10:00"
    if current < SLOT_WINDOWS["16:00"][0]:
        return "10:00"
    if current < SLOT_WINDOWS["20:00"][0]:
        return "16:00"
    return "20:00"


def get_submission_status(time_slot: str, message_time: str) -> str:
    start, end = SLOT_WINDOWS[time_slot]
    if start <= message_time <= end:
        return "on_time"
    return "late"


def parse_time_from_text(text_upper: str) -> str | None:
    match = re.search(r"(\d{1,2})[:.](\d{2})", text_upper)
    if match:
        return f"{int(match.group(1)):02d}:{match.group(2)}"
    for hour in ("10", "16", "20", "9", "11", "12", "13", "14", "15", "17", "18", "19", "21"):
        if f" {hour} " in f" {text_upper} " or text_upper.endswith(f" {hour}"):
            return f"{int(hour):02d}:00"
    return None


def find_point(text_upper: str) -> str | None:
    """Точка как отдельное слово (не подстрока в другом тексте)."""
    ordered = sorted(POINTS, key=len, reverse=True)
    for point in ordered:
        if re.search(
            rf"(?:^|[\s,.:;!\-(]){re.escape(point.upper())}(?:[\s,.:;!\-)($]|$)",
            text_upper,
        ):
            return point
    return None


def has_vitrina_keyword(text_upper: str) -> bool:
    return bool(re.search(r"ВИТРИН", text_upper))


def parse_vitrina_time_slot(text_upper: str) -> str | None:
    """Время слота витрины: только 10:00, 16:00 или 20:00."""
    match = re.search(r"(?:^|[\s,.:;!\-(])(10|16|20)(?:[:.]00)?(?:[\s,.:;!\-)($]|$)", text_upper)
    if match:
        return f"{match.group(1)}:00"
    return None


def normalize_caption(text: str) -> str:
    text = text.strip()
    text = re.sub(r"^@\w+\s*", "", text)
    return text.strip()


def is_vitrina_format(text_upper: str, point: str) -> bool:
    """
    Допустимые форматы:
    - МН 10:00 / МН 10
    - МН витрина 10:00
    Сообщение должно начинаться с кода точки.
    """
    compact = re.sub(r"\s+", " ", text_upper).strip()
    prefix = point.upper()
    if not compact.startswith(prefix):
        return False

    rest = compact[len(prefix):].strip()
    time_part = r"(?:10|16|20)(?:[:.]00)?"

    if re.fullmatch(time_part, rest):
        return True
    if re.match(rf"^ВИТРИН\w*\s+{time_part}$", rest):
        return True
    return False


def week_dates(reference: datetime | None = None) -> list[str]:
    ref = (reference or now_msk()).date()
    monday = ref - timedelta(days=ref.weekday())
    return [(monday + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]


def format_period(dates: list[str]) -> str:
    dates = sorted(dates)
    if len(dates) == 1:
        return dates[0]
    return f"{dates[0]} — {dates[-1]} ({len(dates)} дн.)"


# ====================== БАЗА ======================
def setup_db_location() -> None:
    db_dir = os.path.dirname(os.path.abspath(DB_PATH))
    if db_dir:
        os.makedirs(db_dir, exist_ok=True)
    os.makedirs(BACKUP_DIR, exist_ok=True)


async def import_legacy_db_if_needed() -> None:
    legacy = os.path.abspath(LEGACY_DB_PATH)
    target = os.path.abspath(DB_PATH)
    if legacy == target or not os.path.isfile(legacy):
        return
    if os.path.isfile(target):
        async with aiosqlite.connect(target) as db:
            cursor = await db.execute("SELECT COUNT(*) FROM vitrina_reports")
            row = await cursor.fetchone()
            if row and row[0]:
                return
    shutil.copy2(legacy, target)
    logger.info("База перенесена: %s -> %s", LEGACY_DB_PATH, DB_PATH)


async def fetch_db_summary() -> str:
    if not os.path.isfile(DB_PATH):
        return "База ещё не создана."

    size_kb = os.path.getsize(DB_PATH) // 1024
    async with aiosqlite.connect(DB_PATH) as db:
        cursor = await db.execute(
            "SELECT date, COUNT(*) FROM vitrina_reports GROUP BY date ORDER BY date"
        )
        vitrina_days = await cursor.fetchall()
        cursor = await db.execute("SELECT COUNT(*) FROM vitrina_reports")
        vitrina_total = (await cursor.fetchone())[0]
        cursor = await db.execute(
            "SELECT date, COUNT(*) FROM shift_events GROUP BY date ORDER BY date"
        )
        shift_days = await cursor.fetchall()
        cursor = await db.execute("SELECT COUNT(*) FROM shift_events")
        shift_total = (await cursor.fetchone())[0]

    vitrina_dates = ", ".join(d for d, _ in vitrina_days) or "—"
    shift_dates = ", ".join(d for d, _ in shift_days) or "—"
    return (
        f"Файл: `{DB_PATH}` ({size_kb} KB)\n"
        f"Витрины: {vitrina_total} записей ({len(vitrina_days)} дн.)\n"
        f"Даты витрин: {vitrina_dates}\n"
        f"Выходы: {shift_total} записей ({len(shift_days)} дн.)\n"
        f"Даты выходов: {shift_dates}"
    )


def prune_local_backups(keep: int = BACKUP_KEEP_LOCAL) -> None:
    if not os.path.isdir(BACKUP_DIR):
        return
    files = sorted(
        (
            os.path.join(BACKUP_DIR, name)
            for name in os.listdir(BACKUP_DIR)
            if name.endswith(".db")
        ),
        key=os.path.getmtime,
        reverse=True,
    )
    for path in files[keep:]:
        try:
            os.remove(path)
        except OSError:
            logger.exception("Не удалось удалить старый бэкап: %s", path)


async def create_db_backup_file(prefix: str = "vitrina_backup") -> str | None:
    if not os.path.isfile(DB_PATH):
        return None
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp = now_msk().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"{prefix}_{stamp}.db")
    async with aiosqlite.connect(DB_PATH) as src:
        async with aiosqlite.connect(backup_path) as dst:
            await src.backup(dst)
    prune_local_backups()
    return backup_path


async def backup_database_to_admins(reason: str) -> bool:
    backup_path = await create_db_backup_file()
    if not backup_path:
        logger.warning("Бэкап пропущен: файл базы не найден (%s)", DB_PATH)
        return False

    summary = await fetch_db_summary()
    caption = f"💾 **Бэкап базы** ({reason})\n\n{summary}"
    admin_ids = await fetch_admin_ids()
    if not admin_ids:
        logger.warning("Бэкап создан локально, но нет админов для отправки")
        return True

    sent = 0
    for user_id in admin_ids:
        try:
            await send_admin_dm(
                user_id,
                caption,
                parse_mode="Markdown",
            )
            await bot.send_document(
                user_id,
                FSInputFile(backup_path),
                caption=f"Файл: {os.path.basename(backup_path)}",
            )
            sent += 1
        except Exception:
            logger.exception("Не удалось отправить бэкап админу %s", user_id)

    logger.info("Бэкап (%s): %s, отправлено %d админам", reason, backup_path, sent)
    return True


async def restore_database_from_file(source_path: str) -> str:
    async with aiosqlite.connect(source_path) as db:
        cursor = await db.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='vitrina_reports'"
        )
        if not await cursor.fetchone():
            raise ValueError("В файле нет таблицы vitrina_reports")

    setup_db_location()
    if os.path.isfile(DB_PATH):
        await create_db_backup_file(prefix="before_restore")

    shutil.copy2(source_path, DB_PATH)
    await init_db()
    return await fetch_db_summary()


_last_daily_backup_date: str | None = None
_last_interval_backup_key: str | None = None


async def maybe_run_scheduled_backup(now: datetime) -> None:
    global _last_daily_backup_date, _last_interval_backup_key

    today = now.strftime("%Y-%m-%d")
    if now.hour == 23 and now.minute >= 55 and _last_daily_backup_date != today:
        _last_daily_backup_date = today
        await backup_database_to_admins("ежедневный")

    interval_key = f"{today}-{now.hour // 6}"
    if now.minute == 0 and _last_interval_backup_key != interval_key:
        _last_interval_backup_key = interval_key
        await create_db_backup_file(prefix="auto")


async def init_db() -> None:
    setup_db_location()
    await import_legacy_db_if_needed()
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            """CREATE TABLE IF NOT EXISTS vitrina_reports (
            id INTEGER PRIMARY KEY,
            date TEXT,
            point TEXT,
            time_slot TEXT,
            user_id INTEGER,
            username TEXT,
            message_time TEXT,
            status TEXT,
            has_photo INTEGER DEFAULT 0
        )"""
        )
        await db.execute(
            """CREATE TABLE IF NOT EXISTS shift_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            point TEXT,
            user_id INTEGER,
            username TEXT,
            event_type TEXT,
            event_time TEXT,
            event_label TEXT,
            created_at TEXT
        )"""
        )
        await db.execute(
            """CREATE TABLE IF NOT EXISTS slot_notifications (
            date TEXT,
            time_slot TEXT,
            PRIMARY KEY (date, time_slot)
        )"""
        )
        await db.execute(
            """CREATE TABLE IF NOT EXISTS admins (
            user_id INTEGER PRIMARY KEY,
            username TEXT,
            added_at TEXT
        )"""
        )
        await db.execute(
            "CREATE INDEX IF NOT EXISTS idx_reports_date ON vitrina_reports(date)"
        )
        await db.execute(
            "CREATE INDEX IF NOT EXISTS idx_shifts_date ON shift_events(date)"
        )
        try:
            await db.execute(
                "ALTER TABLE vitrina_reports ADD COLUMN has_photo INTEGER DEFAULT 0"
            )
        except aiosqlite.OperationalError:
            pass
        await migrate_shift_events(db)
        await db.commit()


async def migrate_shift_events(db: aiosqlite.Connection) -> None:
    cursor = await db.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='shift_events'"
    )
    row = await cursor.fetchone()
    if not row or not row[0]:
        return

    table_sql = row[0]
    if "UNIQUE(date, point, user_id, event_type)" not in table_sql:
        try:
            await db.execute(
                "ALTER TABLE shift_events ADD COLUMN event_label TEXT DEFAULT ''"
            )
        except aiosqlite.OperationalError:
            pass
        try:
            await db.execute(
                "ALTER TABLE shift_events ADD COLUMN created_at TEXT DEFAULT ''"
            )
        except aiosqlite.OperationalError:
            pass
        return

    await db.execute(
        """CREATE TABLE IF NOT EXISTS shift_events_new (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            point TEXT,
            user_id INTEGER,
            username TEXT,
            event_type TEXT,
            event_time TEXT,
            event_label TEXT,
            created_at TEXT
        )"""
    )
    await db.execute(
        """INSERT INTO shift_events_new
           (date, point, user_id, username, event_type, event_time, event_label, created_at)
           SELECT date, point, user_id, username, event_type, event_time,
                  CASE WHEN event_type = 'arrival' THEN 'пришёл' ELSE 'ушёл' END,
                  date || ' ' || event_time
           FROM shift_events"""
    )
    await db.execute("DROP TABLE shift_events")
    await db.execute("ALTER TABLE shift_events_new RENAME TO shift_events")
    await db.execute(
        "CREATE INDEX IF NOT EXISTS idx_shifts_date ON shift_events(date)"
    )


async def fetch_reports_for_dates(dates: list[str]) -> dict[tuple[str, str, str], dict]:
    if not dates:
        return {}
    placeholders = ",".join("?" * len(dates))
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute(
            f"""SELECT date, point, time_slot, username, message_time, status, has_photo
                FROM vitrina_reports WHERE date IN ({placeholders})""",
            dates,
        )
        rows = await cursor.fetchall()
    return {(row["date"], row["point"], row["time_slot"]): dict(row) for row in rows}


async def fetch_shifts_for_dates(dates: list[str]) -> list[dict]:
    if not dates:
        return []
    placeholders = ",".join("?" * len(dates))
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute(
            f"""SELECT date, point, user_id, username, event_type, event_time, event_label
                FROM shift_events WHERE date IN ({placeholders})
                ORDER BY date, point, created_at, id, event_time""",
            dates,
        )
        rows = await cursor.fetchall()
    return [dict(row) for row in rows]


async def is_slot_notified(date: str, time_slot: str) -> bool:
    async with aiosqlite.connect(DB_PATH) as db:
        cursor = await db.execute(
            "SELECT 1 FROM slot_notifications WHERE date=? AND time_slot=?",
            (date, time_slot),
        )
        return await cursor.fetchone() is not None


async def mark_slot_notified(date: str, time_slot: str) -> None:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT OR IGNORE INTO slot_notifications (date, time_slot) VALUES (?, ?)",
            (date, time_slot),
        )
        await db.commit()


async def load_admins() -> set[int]:
    async with aiosqlite.connect(DB_PATH) as db:
        cursor = await db.execute("SELECT user_id FROM admins")
        rows = await cursor.fetchall()
    return {row[0] for row in rows}


async def save_admin(user_id: int, username: str | None) -> None:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            """INSERT INTO admins (user_id, username, added_at)
               VALUES (?, ?, ?)
               ON CONFLICT(user_id) DO UPDATE SET username=excluded.username""",
            (user_id, username or "", now_msk().strftime("%Y-%m-%d %H:%M")),
        )
        await db.commit()


async def fetch_admin_ids() -> list[int]:
    async with aiosqlite.connect(DB_PATH) as db:
        cursor = await db.execute("SELECT user_id FROM admins")
        rows = await cursor.fetchall()
    return [row[0] for row in rows]


async def send_admin_dm(user_id: int, text: str, **kwargs) -> None:
    """Уведомления только в ЛС. В рабочую группу не отправляем."""
    if user_id == WORK_CHAT_ID or user_id < 0:
        logger.warning("Пропущена отправка в чат %s (не ЛС админа)", user_id)
        return
    await bot.send_message(user_id, text, **kwargs)


async def notify_admins(text: str) -> None:
    admin_ids = await fetch_admin_ids()
    if not admin_ids:
        logger.warning("Некому отправить уведомление — нет админов в базе")
        return

    for user_id in admin_ids:
        try:
            await send_admin_dm(user_id, text, parse_mode="Markdown")
        except Exception:
            logger.exception("Не удалось отправить уведомление админу %s", user_id)


async def notify_admins_shift(
    point: str,
    event: str,
    event_time: str,
    username: str,
    event_label: str,
) -> None:
    kind = "Приход" if event == "arrival" else "Уход"
    text = (
        f"👥 **{kind} на точке {point}** ({event_label})\n"
        f"Сотрудник: {username}\n"
        f"Время: {event_time}\n"
        f"Дата: {get_today()}"
    )
    await notify_admins(text)


async def notify_admins_vitrina(
    point: str,
    time_slot: str,
    username: str,
    message_time: str,
    status: str,
    has_photo: bool,
) -> None:
    if not has_photo:
        text = (
            f"⚠️ **Витрина {point}** ({time_slot})\n"
            f"Сотрудник: {username}\n"
            f"Время: {message_time}\n"
            f"Статус: нет фото"
        )
    else:
        st = "вовремя" if status == "on_time" else "опоздание"
        text = (
            f"📸 **Витрина {point}** ({time_slot})\n"
            f"Сотрудник: {username}\n"
            f"Время: {message_time}\n"
            f"Статус: {st}"
        )
    await notify_admins(text)


# ====================== ПАРСЕР ======================
def build_message_text_with_reply(message: Message) -> str:
    text = (message.text or message.caption or "").strip()
    if not message.reply_to_message:
        return text
    reply = message.reply_to_message
    reply_text = (reply.text or reply.caption or "").strip()
    if not reply_text:
        return text
    return f"{reply_text} {text}".strip()


def classify_shift_event(text_upper: str) -> tuple[str | None, str | None]:
    if any(k in text_upper for k in RETURN_KEYWORDS):
        return "arrival", "вернулся"
    if any(k in text_upper for k in STEP_OUT_KEYWORDS):
        return "departure", "отошёл"
    if any(k in text_upper for k in ARRIVAL_KEYWORDS):
        return "arrival", "пришёл"
    if any(k in text_upper for k in DEPARTURE_KEYWORDS):
        return "departure", "ушёл"
    return None, None


def parse_shift_message(text: str) -> tuple[str | None, str | None, str | None, str | None]:
    text_upper = normalize_caption(text).upper()
    event, event_label = classify_shift_event(text_upper)
    if not event:
        return None, None, None, None

    point = find_point(text_upper)
    if not point:
        return None, None, None, None

    event_time = parse_time_from_text(text_upper) or now_msk().strftime("%H:%M")
    return point, event, event_time, event_label


async def save_shift_event(
    date: str,
    point: str,
    user_id: int,
    username: str,
    event: str,
    event_time: str,
    event_label: str,
) -> None:
    created_at = now_msk().strftime("%Y-%m-%d %H:%M:%S")
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            """INSERT INTO shift_events
               (date, point, user_id, username, event_type, event_time, event_label, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (date, point, user_id, username, event, event_time, event_label, created_at),
        )
        await db.commit()


def collect_shift_event_lines(
    dates: list[str],
    shifts: list[dict],
) -> list[tuple[str, str, str, str, str]]:
    """date, point, username, event_label, event_time"""
    lines: list[tuple[str, str, str, str, str]] = []
    for date in sorted(dates):
        day_events = [s for s in shifts if s["date"] == date]
        by_point: dict[str, list[dict]] = {}
        for ev in day_events:
            by_point.setdefault(ev["point"], []).append(ev)

        for point in POINTS:
            for ev in by_point.get(point, []):
                label = ev.get("event_label") or (
                    "пришёл" if ev["event_type"] == "arrival" else "ушёл"
                )
                lines.append((
                    date,
                    point,
                    ev["username"],
                    label,
                    ev["event_time"],
                ))
    return lines


def build_shift_text_blocks(
    dates: list[str],
    shifts: list[dict],
) -> list[str]:
    lines = collect_shift_event_lines(dates, shifts)
    if not lines:
        return ["\n👥 **ВЫХОДЫ СОТРУДНИКОВ**\n\nНет данных о выходах"]

    blocks = ["\n👥 **ВЫХОДЫ СОТРУДНИКОВ**"]
    by_date: dict[str, list[tuple[str, str, str, str, str]]] = {}
    for row in lines:
        by_date.setdefault(row[0], []).append(row)

    for date in sorted(by_date.keys()):
        blocks.append(f"\n📅 **{date}**")
        date_rows = by_date[date]
        for point in POINTS:
            point_rows = [r for r in date_rows if r[1] == point]
            if not point_rows:
                continue
            point_lines = [f"**Точка: {point}**"]
            for _, _, username, label, event_time in point_rows:
                point_lines.append(f"  • {username} — {label} {event_time}")
            blocks.append("\n".join(point_lines))

    return blocks


def parse_vitrina_message(text: str) -> tuple[str | None, str | None]:
    text_upper = normalize_caption(text).upper()
    if not text_upper:
        return None, None

    if any(k in text_upper for k in ARRIVAL_KEYWORDS + DEPARTURE_KEYWORDS):
        return None, None

    point = find_point(text_upper)
    if not point:
        return None, None

    time_slot = parse_vitrina_time_slot(text_upper)
    if not time_slot:
        return None, None

    if not is_vitrina_format(text_upper, point):
        return None, None

    if not is_vitrina_required(get_today(), point, time_slot):
        return point, None

    return point, time_slot


# ====================== СТАТИСТИКА ======================
def count_required_reports(dates: list[str]) -> int:
    total = 0
    for date in dates:
        for point in POINTS:
            total += len(required_slots_for_date(date, point))
    return total


def compute_point_rating(
    dates: list[str],
    reports: dict[tuple[str, str, str], dict],
) -> list[tuple[str, int]]:
    misses: dict[str, int] = {p: 0 for p in POINTS}
    for date in dates:
        for point in POINTS:
            for slot in required_slots_for_date(date, point):
                report = reports.get((date, point, slot))
                if not report or not report.get("has_photo"):
                    misses[point] += 1
    return sorted(misses.items(), key=lambda x: (-x[1], x[0]))


def compute_fines(
    dates: list[str],
    reports: dict[tuple[str, str, str], dict],
) -> int:
    fines = 0
    for date in dates:
        for point in POINTS:
            for slot in required_slots_for_date(date, point):
                report = reports.get((date, point, slot))
                if not report or not report.get("has_photo"):
                    fines += FINE_AMOUNT
                elif report.get("status") == "late":
                    fines += FINE_AMOUNT
    return fines


def status_label(status: str) -> str:
    return "✓ вовремя" if status == "on_time" else "⚠ опоздание"


def format_vitrina_slot_line(slot: str, report: dict | None) -> str:
    if report and report.get("has_photo"):
        icon = "✅" if report["status"] == "on_time" else "⚠️"
        st = "вовремя" if report["status"] == "on_time" else "опоздание"
        return (
            f"  • **{slot}** — {icon} {report['username']} "
            f"({report['message_time']}, {st})"
        )
    return f"  • **{slot}** — ❌ нет фото"


def build_vitrina_text_blocks(
    dates: list[str],
    reports: dict[tuple[str, str, str], dict],
) -> list[str]:
    blocks: list[str] = []
    for date in dates:
        blocks.append(f"\n📅 **{date}**")
        point_blocks: list[str] = []
        for point in POINTS:
            slots = required_slots_for_date(date, point)
            lines = [f"**Точка: {point}**"]
            for slot in slots:
                report = reports.get((date, point, slot))
                lines.append(format_vitrina_slot_line(slot, report))
            point_blocks.append("\n".join(lines))
        blocks.append("\n\n".join(point_blocks))
    return blocks


def is_admin(user_id: int) -> bool:
    return user_id in authenticated_admins


def build_admin_keyboard(manual_mode: bool = False) -> ReplyKeyboardMarkup:
    rows = [
        [KeyboardButton(text="📊 Отчёт")],
        [KeyboardButton(text="📋 Статус сегодня")],
        [KeyboardButton(text="✏️ Внести витрину")],
        [KeyboardButton(text="💾 Бэкап базы")],
    ]
    if manual_mode:
        rows.append([KeyboardButton(text="❌ Отмена")])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def manual_vitrina_help_text() -> str:
    return (
        "✏️ **Ручной ввод витрины**\n\n"
        "Отправьте **одной строкой**:\n"
        "`дата точка время имя`\n\n"
        "**Примеры:**\n"
        "• `2026-06-15 МН 10:00 Иван`\n"
        "• `15.06.2026 СМ 16 Мария`\n"
        "• `2026-06-15 ПТ 20:00 @ivanov`\n\n"
        "Если запись за этот день/точку/слот уже есть — она **заменится**.\n"
        "Для отмены нажмите **❌ Отмена**."
    )


def parse_manual_vitrina(text: str) -> tuple[str, str, str, str] | None:
    """date, point, time_slot, username"""
    text = text.strip()
    match = re.match(
        r"^(?:(\d{4}-\d{2}-\d{2})|(\d{1,2})\.(\d{1,2})\.(\d{4}))\s+"
        r"(.+?)\s+"
        r"(\d{1,2})(?::(\d{2}))?\s+"
        r"(.+)$",
        text,
        re.IGNORECASE,
    )
    if not match:
        return None

    if match.group(1):
        date = match.group(1)
    else:
        date = (
            f"{match.group(4)}-{int(match.group(3)):02d}-{int(match.group(2)):02d}"
        )

    point_raw = match.group(5).upper()
    point = find_point(point_raw) or find_point(f" {point_raw} ")
    if not point:
        for p in POINTS:
            if point_raw == p.upper():
                point = p
                break
    if not point:
        return None

    hour = int(match.group(6))
    slot = f"{hour:02d}:00"
    if slot not in TIMES:
        return None
    if not is_vitrina_required(date, point, slot):
        return None

    username = match.group(8).strip().lstrip("@")
    if not username:
        return None

    return date, point, slot, username


async def save_manual_vitrina(
    date: str,
    point: str,
    time_slot: str,
    username: str,
) -> None:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "DELETE FROM vitrina_reports WHERE date=? AND point=? AND time_slot=?",
            (date, point, time_slot),
        )
        await db.execute(
            """INSERT INTO vitrina_reports
               (date, point, time_slot, user_id, username, message_time, status, has_photo)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (date, point, time_slot, 0, username, time_slot, "on_time", 1),
        )
        await db.commit()


def admin_welcome_text() -> str:
    return (
        "🔐 **Админ-панель**\n\n"
        "Выберите действие кнопкой ниже.\n"
        "Отчёты приходят сюда, в личные сообщения.\n\n"
        "💾 Бэкап базы — копия файла сюда в ЛС.\n"
        "Чтобы восстановить базу — отправьте файл `.db` в этот чат."
    )


async def build_today_status_text() -> str:
    today = get_today()
    reports = await fetch_reports_for_dates([today])
    lines = [f"📋 **Статус на сегодня** ({today})\n"]
    point_blocks: list[str] = []

    for point in POINTS:
        block_lines = [f"**Точка: {point}**"]
        for slot in required_slots_for_date(today, point):
            report = reports.get((today, point, slot))
            block_lines.append(format_vitrina_slot_line(slot, report))
        point_blocks.append("\n".join(block_lines))

    lines.append("\n\n".join(point_blocks))
    return "\n".join(lines)


async def send_report_calendar(message: Message, user_id: int) -> None:
    session = get_admin_session(user_id)
    session["dates"] = set()
    await message.answer(
        calendar_caption(session),
        parse_mode="Markdown",
        reply_markup=build_calendar(session["year"], session["month"], session["dates"]),
    )


# ====================== КАЛЕНДАРЬ ======================
def get_admin_session(user_id: int) -> dict:
    if user_id not in admin_sessions:
        now = now_msk()
        admin_sessions[user_id] = {
            "dates": set(),
            "year": now.year,
            "month": now.month,
            "mode": None,
        }
    return admin_sessions[user_id]


def build_calendar(
    year: int,
    month: int,
    selected: set[str] | None = None,
) -> InlineKeyboardMarkup:
    selected = selected or set()
    today = now_msk().date()
    cal = calendar.Calendar(firstweekday=0)
    weeks = cal.monthdayscalendar(year, month)

    buttons: list[list[InlineKeyboardButton]] = [
        [
            InlineKeyboardButton(
                text=f"◀ {MONTH_NAMES[month - 1] if month > 1 else MONTH_NAMES[12]}",
                callback_data=f"cal:prev:{year}:{month}",
            ),
            InlineKeyboardButton(
                text=f"{MONTH_NAMES[month]} {year}",
                callback_data="cal:noop",
            ),
            InlineKeyboardButton(
                text=f"{MONTH_NAMES[month + 1] if month < 12 else MONTH_NAMES[1]} ▶",
                callback_data=f"cal:next:{year}:{month}",
            ),
        ],
        [
            InlineKeyboardButton(text="Пн", callback_data="cal:noop"),
            InlineKeyboardButton(text="Вт", callback_data="cal:noop"),
            InlineKeyboardButton(text="Ср", callback_data="cal:noop"),
            InlineKeyboardButton(text="Чт", callback_data="cal:noop"),
            InlineKeyboardButton(text="Пт", callback_data="cal:noop"),
            InlineKeyboardButton(text="Сб", callback_data="cal:noop"),
            InlineKeyboardButton(text="Вс", callback_data="cal:noop"),
        ],
    ]

    for week in weeks:
        row: list[InlineKeyboardButton] = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(text=" ", callback_data="cal:noop"))
                continue
            date_str = f"{year:04d}-{month:02d}-{day:02d}"
            label = str(day)
            if date_str in selected:
                label = f"✓{day}"
            elif year == today.year and month == today.month and day == today.day:
                label = f"•{day}•"
            row.append(
                InlineKeyboardButton(text=label, callback_data=f"cal:toggle:{date_str}")
            )
        buttons.append(row)

    buttons.append([
        InlineKeyboardButton(text="📅 Неделя", callback_data="cal:week"),
        InlineKeyboardButton(text="🗑 Сброс", callback_data="cal:clear"),
    ])
    buttons.append([
        InlineKeyboardButton(text="✅ Готово", callback_data="cal:done"),
        InlineKeyboardButton(text="◀ Меню", callback_data="adm:menu"),
    ])

    return InlineKeyboardMarkup(inline_keyboard=buttons)


def build_format_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📊 Excel", callback_data="cal:fmt:xlsx"),
            InlineKeyboardButton(text="📝 Текст", callback_data="cal:fmt:txt"),
        ],
    ])


def calendar_caption(session: dict) -> str:
    dates = sorted(session["dates"])
    if dates:
        dates_text = f"Выбрано: {format_period(dates)}"
    else:
        dates_text = "Нажмите на даты (можно несколько) или «📅 Неделя»"
    return (
        "📅 **Выбор дат для отчёта**\n\n"
        f"{dates_text}\n\n"
        "Отчёт включает:\n"
        "• витрины (фото)\n"
        "• выходы сотрудников (приход/уход)\n"
        "• рейтинг точек по пропускам"
    )


# ====================== ОТЧЁТЫ ======================
def build_text_report(
    dates: list[str],
    reports: dict[tuple[str, str, str], dict],
    shifts: list[dict],
) -> list[str]:
    dates = sorted(dates)
    rating = compute_point_rating(dates, reports)
    fines = compute_fines(dates, reports)
    chunks: list[str] = []

    header = (
        f"📊 **ОТЧЁТ** {format_period(dates)}\n"
        f"Депремия: {fines} руб.\n"
        f"{'=' * 30}"
    )
    chunks.append(header)

    rating_lines = ["\n🏆 **РЕЙТИНГ ТОЧЕК** (пропуски фото):"]
    if any(count for _, count in rating):
        for i, (point, count) in enumerate(rating, 1):
            if count:
                rating_lines.append(f"{i}. **{point}** — {count} пропуск(ов)")
    else:
        rating_lines.append("Пропусков нет ✓")
    chunks.append("\n".join(rating_lines))

    vitrina_header = "\n📸 **ВИТРИНЫ**"
    chunks.append(vitrina_header)
    chunks.extend(build_vitrina_text_blocks(dates, reports))
    chunks.extend(build_shift_text_blocks(dates, shifts))

    return split_text_chunks(chunks)


def split_text_chunks(parts: list[str], limit: int = 4000) -> list[str]:
    messages: list[str] = []
    current = ""
    for part in parts:
        if len(current) + len(part) + 1 > limit and current:
            messages.append(current.strip())
            current = part
        else:
            current = f"{current}\n{part}" if current else part
    if current.strip():
        messages.append(current.strip())
    return messages or ["Нет данных за выбранный период."]


async def generate_report_xlsx(
    dates: list[str],
    reports: dict[tuple[str, str, str], dict],
    shifts: list[dict],
) -> str:
    dates = sorted(dates)
    rating = compute_point_rating(dates, reports)
    fines = compute_fines(dates, reports)

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    late_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # --- Лист: витрины ---
    ws = wb.active
    ws.title = "Витрины"
    ws.append(["Период:", format_period(dates)])
    ws.append(["Депремия:", fines, "руб."])
    ws.append([])

    header = ["Дата", "Точка"]
    for slot in TIMES:
        header.extend([f"{slot} — сотрудник", f"{slot} — время", f"{slot} — статус"])
    ws.append(header)
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font

    for date in dates:
        for point in POINTS:
            row = [date, point]
            for slot in TIMES:
                if not is_vitrina_required(date, point, slot):
                    row.extend(["—", "—", "не требуется"])
                    continue
                report = reports.get((date, point, slot))
                if report and report.get("has_photo"):
                    row.extend([
                        report["username"],
                        report["message_time"],
                        status_label(report["status"]),
                    ])
                else:
                    row.extend(["—", "—", "нет фото"])
            ws.append(row)
            row_idx = ws.max_row
            for col_idx, value in enumerate(row[2:], start=3):
                if value == "нет фото":
                    ws.cell(row=row_idx, column=col_idx).fill = missing_fill
                elif value == "⚠ опоздание":
                    ws.cell(row=row_idx, column=col_idx).fill = late_fill

    # --- Лист: выходы ---
    ws2 = wb.create_sheet("Выходы")
    ws2.append(["Дата", "Точка", "Сотрудник", "Событие", "Время"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    for date, point, username, label, event_time in collect_shift_event_lines(dates, shifts):
        ws2.append([date, point, username, label, event_time])

    if ws2.max_row == 1:
        ws2.append(["—", "—", "Нет данных", "—", "—"])

    # --- Лист: рейтинг ---
    ws3 = wb.create_sheet("Рейтинг")
    ws3.append(["Место", "Точка", "Пропусков фото"])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    for i, (point, count) in enumerate(rating, 1):
        ws3.append([i, point, count])

    for sheet in (ws, ws2, ws3):
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    period = dates[0] if len(dates) == 1 else f"{dates[0]}_{dates[-1]}"
    filename = f"report_{period}.xlsx"
    wb.save(filename)
    return filename


# ====================== УВЕДОМЛЕНИЯ (только ЛС админам, не в группу) ======================
async def check_missed_vitrinas() -> None:
    now = now_msk()
    today = now.strftime("%Y-%m-%d")
    current = now.strftime("%H:%M")

    for slot, end_time in SLOT_END.items():
        if current <= end_time:
            continue
        if await is_slot_notified(today, slot):
            continue

        reports = await fetch_reports_for_dates([today])
        missed: list[str] = []
        for point in POINTS:
            if not is_vitrina_required(today, point, slot):
                continue
            report = reports.get((today, point, slot))
            if not report or not report.get("has_photo"):
                missed.append(point)

        if missed:
            points_text = ", ".join(f"**{p}**" for p in missed)
            text = (
                f"⚠️ **Пропущены витрины** ({slot}, {today}):\n"
                f"{points_text}\n\n"
                f"Депремия: {FINE_AMOUNT} руб. за каждую точку"
            )
            await notify_admins(text)

        await mark_slot_notified(today, slot)


async def notification_scheduler() -> None:
    while True:
        try:
            now = now_msk()
            await check_missed_vitrinas()
            await maybe_run_scheduled_backup(now)
        except Exception:
            logger.exception("Ошибка в планировщике уведомлений")
        await asyncio.sleep(60)


# ====================== ОБРАБОТКА: РАБОЧИЙ ЧАТ ======================
class WorkChatFilter(BaseFilter):
    async def __call__(self, message: Message) -> bool:
        return message.chat.id == WORK_CHAT_ID


@dp.message(
    WorkChatFilter(),
    (F.text & ~F.text.startswith("/")) | (F.caption & ~F.caption.startswith("/")),
)
async def handle_work_chat_message(message: Message):
    text = build_message_text_with_reply(message)
    if not text or text.split()[0] == ADMIN_CODE:
        return

    has_photo = bool(message.photo)

    point, event, event_time, event_label = parse_shift_message(text)
    if point and event:
        user_id = message.from_user.id
        username = message.from_user.username or message.from_user.full_name
        date = get_today()
        await save_shift_event(
            date, point, user_id, username, event, event_time, event_label or "",
        )

        logger.info(
            "Выход: %s %s %s %s (%s) user=%s reply=%s",
            date, point, event, event_label, event_time, user_id,
            bool(message.reply_to_message),
        )
        await notify_admins_shift(
            point, event, event_time, username, event_label or "",
        )
        return

    point, time_slot = parse_vitrina_message(text)
    if not point:
        return

    if time_slot is None:
        logger.info("Вечерняя витрина отклонена: %s", point)
        return

    user_id = message.from_user.id
    username = message.from_user.username or message.from_user.full_name
    date = get_today()
    message_time = now_msk().strftime("%H:%M")
    status = get_submission_status(time_slot, message_time)

    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "DELETE FROM vitrina_reports WHERE date=? AND point=? AND time_slot=?",
            (date, point, time_slot),
        )
        await db.execute(
            """INSERT INTO vitrina_reports
               (date, point, time_slot, user_id, username, message_time, status, has_photo)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (date, point, time_slot, user_id, username, message_time, status, int(has_photo)),
        )
        await db.commit()

    logger.info(
        "Витрина: %s %s %s photo=%s user=%s",
        date, point, time_slot, has_photo, user_id,
    )
    await notify_admins_vitrina(
        point, time_slot, username, message_time, status, has_photo,
    )


# ====================== КОМАНДЫ ======================
@dp.message(Command("chatid", "chat_id", "чатайди"))
async def cmd_chatid(message: Message):
    if message.chat.id == WORK_CHAT_ID:
        return

    chat = message.chat
    type_names = {
        "private": "личные сообщения",
        "group": "группа",
        "supergroup": "супергруппа",
        "channel": "канал",
    }
    type_name = type_names.get(chat.type, chat.type)
    title = chat.title or chat.full_name or "—"

    text = (
        f"**ID чата:** `{chat.id}`\n"
        f"**Тип:** {type_name}\n"
        f"**Название:** {title}"
    )
    if chat.type in ("group", "supergroup"):
        match = "✅ Это рабочий чат бота" if chat.id == WORK_CHAT_ID else "⚠️ Это не рабочий чат бота"
        text += f"\n\n{match}\n\nДобавьте в `.env`:\n`WORK_CHAT_ID={chat.id}`"

    await message.reply(text, parse_mode="Markdown")


# ====================== ОБРАБОТКА: ЛС АДМИН ======================
@dp.message(Command("start"), F.chat.type == "private")
async def cmd_start(message: Message):
    if is_admin(message.from_user.id):
        await message.answer(
            admin_welcome_text(),
            parse_mode="Markdown",
            reply_markup=build_admin_keyboard(),
        )
        return
    await message.answer(
        "👋 Бот учёта витрин.\n\n"
        "Сотрудники отправляют фото и выходы с точек в **групповом чате**.\n\n"
        "Для доступа к админ-панели введите код.",
        parse_mode="Markdown",
    )


@dp.message(F.chat.type == "private", F.text)
async def handle_private_message(message: Message):
    text = (message.text or "").strip()
    user_id = message.from_user.id

    if text.startswith("/"):
        return

    if text == ADMIN_CODE:
        username = message.from_user.username or message.from_user.full_name
        authenticated_admins.add(user_id)
        await save_admin(user_id, username)
        await message.answer(
            "✅ Доступ открыт!\n\n"
            + admin_welcome_text()
            + "\n\n_Вы будете получать уведомления о пропусках витрин в ЛС._",
            parse_mode="Markdown",
            reply_markup=build_admin_keyboard(),
        )
        return

    if not is_admin(user_id):
        await message.answer("🔒 Введите код администратора для доступа к панели.")
        return

    session = get_admin_session(user_id)

    if text == "❌ Отмена":
        session["mode"] = None
        await message.answer(
            "Ввод отменён.",
            reply_markup=build_admin_keyboard(),
        )
        return

    if text == "✏️ Внести витрину":
        session["mode"] = "manual_vitrina"
        await message.answer(
            manual_vitrina_help_text(),
            parse_mode="Markdown",
            reply_markup=build_admin_keyboard(manual_mode=True),
        )
        return

    if session.get("mode") == "manual_vitrina":
        parsed = parse_manual_vitrina(text)
        if not parsed:
            await message.answer(
                "Не удалось разобрать строку.\n\n" + manual_vitrina_help_text(),
                parse_mode="Markdown",
                reply_markup=build_admin_keyboard(manual_mode=True),
            )
            return

        date, point, time_slot, username = parsed
        await save_manual_vitrina(date, point, time_slot, username)
        session["mode"] = None
        logger.info(
            "Ручной ввод витрины: %s %s %s %s admin=%s",
            date, point, time_slot, username, user_id,
        )
        await message.answer(
            f"✅ Записано:\n"
            f"**{date}** — **{point}** — **{time_slot}**\n"
            f"Сотрудник: {username}",
            parse_mode="Markdown",
            reply_markup=build_admin_keyboard(),
        )
        return

    if text == "📊 Отчёт":
        await send_report_calendar(message, user_id)
        return

    if text == "📋 Статус сегодня":
        await message.answer(
            await build_today_status_text(),
            parse_mode="Markdown",
            reply_markup=build_admin_keyboard(),
        )
        return

    if text == "💾 Бэкап базы":
        ok = await backup_database_to_admins("по запросу")
        if ok:
            await message.answer(
                "✅ Бэкап создан и отправлен в этот чат.",
                reply_markup=build_admin_keyboard(),
            )
        else:
            await message.answer(
                "⚠️ База пуста или файл не найден.",
                reply_markup=build_admin_keyboard(),
            )
        return

    await message.answer(
        "Используйте кнопки ниже 👇",
        reply_markup=build_admin_keyboard(),
    )


@dp.message(F.chat.type == "private", F.document)
async def handle_admin_db_restore(message: Message):
    if not is_admin(message.from_user.id):
        return

    doc = message.document
    filename = (doc.file_name or "").lower()
    if not filename.endswith(".db"):
        return

    temp_path = os.path.join(BACKUP_DIR, f"upload_{message.from_user.id}_{now_msk().strftime('%Y%m%d_%H%M%S')}.db")
    os.makedirs(BACKUP_DIR, exist_ok=True)
    try:
        tg_file = await bot.get_file(doc.file_id)
        await bot.download_file(tg_file.file_path, destination=temp_path)
        summary = await restore_database_from_file(temp_path)
        await message.answer(
            "✅ **База восстановлена из файла**\n\n" + summary,
            parse_mode="Markdown",
            reply_markup=build_admin_keyboard(),
        )
        logger.info("База восстановлена админом %s из %s", message.from_user.id, filename)
    except Exception as exc:
        logger.exception("Ошибка восстановления базы")
        await message.answer(
            f"❌ Не удалось восстановить базу: {exc}",
            reply_markup=build_admin_keyboard(),
        )
    finally:
        try:
            if os.path.isfile(temp_path):
                os.remove(temp_path)
        except OSError:
            pass


@dp.callback_query(F.data.startswith("adm:"))
async def handle_admin_menu(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("Сначала введите код в ЛС боту", show_alert=True)
        return

    if callback.data.split(":")[1] == "menu":
        try:
            await callback.message.edit_text(admin_welcome_text(), parse_mode="Markdown")
        except Exception:
            pass
        await callback.message.answer(
            "Выберите действие:",
            reply_markup=build_admin_keyboard(),
        )
        await callback.answer()
        return

    await callback.answer()


@dp.callback_query(F.data.startswith("cal:"))
async def handle_calendar(callback: CallbackQuery):
    if callback.message.chat.type != "private" or not is_admin(callback.from_user.id):
        await callback.answer("Доступ только для админов в ЛС", show_alert=True)
        return
    parts = callback.data.split(":")
    action = parts[1]
    user_id = callback.from_user.id
    session = get_admin_session(user_id)

    if action == "noop":
        await callback.answer()
        return

    if action in ("prev", "next"):
        year, month = int(parts[2]), int(parts[3])
        if action == "prev":
            month -= 1
            if month < 1:
                month, year = 12, year - 1
        else:
            month += 1
            if month > 12:
                month, year = 1, year + 1
        session["year"], session["month"] = year, month
        await callback.message.edit_text(
            calendar_caption(session),
            parse_mode="Markdown",
            reply_markup=build_calendar(year, month, session["dates"]),
        )
        await callback.answer()
        return

    if action == "toggle":
        date = parts[2]
        if date in session["dates"]:
            session["dates"].remove(date)
        else:
            session["dates"].add(date)
        await callback.message.edit_text(
            calendar_caption(session),
            parse_mode="Markdown",
            reply_markup=build_calendar(
                session["year"], session["month"], session["dates"]
            ),
        )
        await callback.answer()
        return

    if action == "week":
        session["dates"].update(week_dates())
        await callback.message.edit_text(
            calendar_caption(session),
            parse_mode="Markdown",
            reply_markup=build_calendar(
                session["year"], session["month"], session["dates"]
            ),
        )
        await callback.answer("Выбрана текущая неделя")
        return

    if action == "clear":
        session["dates"].clear()
        await callback.message.edit_text(
            calendar_caption(session),
            parse_mode="Markdown",
            reply_markup=build_calendar(
                session["year"], session["month"], session["dates"]
            ),
        )
        await callback.answer("Выбор сброшен")
        return

    if action == "done":
        if not session["dates"]:
            await callback.answer("Выберите хотя бы одну дату", show_alert=True)
            return
        await callback.message.edit_text(
            f"📋 Период: **{format_period(sorted(session['dates']))}**\n\n"
            "Выберите формат отчёта:",
            parse_mode="Markdown",
            reply_markup=build_format_keyboard(),
        )
        await callback.answer()
        return

    if action == "fmt":
        fmt = parts[2]
        dates = sorted(session["dates"])
        await callback.answer("Формирую отчёт...")
        await callback.message.edit_text(
            f"⏳ Формирую отчёт ({format_period(dates)})..."
        )

        reports = await fetch_reports_for_dates(dates)
        shifts = await fetch_shifts_for_dates(dates)

        if fmt == "xlsx":
            filename = await generate_report_xlsx(dates, reports, shifts)
            await callback.message.answer_document(
                FSInputFile(filename),
                caption=f"📊 Отчёт за {format_period(dates)}",
            )
            try:
                os.remove(filename)
            except OSError:
                pass
        else:
            for chunk in build_text_report(dates, reports, shifts):
                await callback.message.answer(chunk, parse_mode="Markdown")

        session["dates"] = set()
        try:
            await callback.message.edit_text(
                admin_welcome_text(),
                parse_mode="Markdown",
            )
        except Exception:
            pass
        await callback.message.answer(
            "Отчёт готов. Выберите следующее действие:",
            reply_markup=build_admin_keyboard(),
        )
        return

    await callback.answer()


# ====================== ЗАПУСК ======================
def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        force=True,
    )


async def main():
    setup_logging()
    print("=== VITRINA BOT: старт ===", flush=True)
    print(f"=== TOKEN: {'задан' if TOKEN else 'НЕТ'} ===", flush=True)
    print(f"=== WORK_CHAT_ID: {WORK_CHAT_ID} ===", flush=True)
    print(f"=== DB_PATH: {DB_PATH} ===", flush=True)

    await init_db()
    authenticated_admins.update(await load_admins())
    logger.info("Админов в базе: %d", len(authenticated_admins))
    logger.info("База данных: %s", DB_PATH)
    logger.info("Сводка базы:\n%s", await fetch_db_summary())
    await create_db_backup_file(prefix="startup")
    logger.info("Бот запущен. Слоты (МСК): %s", ", ".join(TIMES))
    logger.info("Рабочий чат: chat_id=%s", WORK_CHAT_ID)
    print("=== VITRINA BOT: polling... ===", flush=True)

    asyncio.create_task(notification_scheduler())
    await dp.start_polling(bot)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception:
        print("=== VITRINA BOT: ОШИБКА ЗАПУСКА ===", flush=True)
        logging.exception("Критическая ошибка при запуске")
        raise
